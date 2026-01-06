# src/layout_verifier/excel_colorizer.py
#
# Excel cell coloring based on verification results.
# Uses openpyxl to modify cell background colors while preserving images.

import shutil
import zipfile
import tempfile
from pathlib import Path
from typing import Optional
from dataclasses import dataclass

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from .logging_utils import log_info, log_warning


# Color definitions (RGB hex values)
GREEN_FILL = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green - match
RED_FILL = PatternFill(start_color="FFB6B6", end_color="FFB6B6", fill_type="solid")  # Light red - no match
YELLOW_FILL = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")  # Light yellow - unchecked
NO_FILL = PatternFill(fill_type=None)  # No background - product not found


@dataclass
class CellColorResult:
    """Tracks which cells were colored and how."""
    row: int
    column: str
    color: str  # "green", "red", "yellow", or "none"
    field_name: str
    value: str


@dataclass
class ColoringResult:
    """Summary of Excel coloring operation."""
    excel_path: str
    products_found: int
    products_not_found: int
    cells_green: int  # Matched
    cells_red: int  # Not matched
    cells_yellow: int  # Unchecked
    cell_details: list[CellColorResult]


def find_column_indices(ws, target_columns: list[str]) -> dict[str, int]:
    """
    Find column indices for target columns in the worksheet.

    Args:
        ws: openpyxl worksheet
        target_columns: List of column names to find

    Returns:
        Dictionary mapping column name to column index (1-based)
    """
    column_map = {}
    header_row = 1

    for col_idx, cell in enumerate(ws[header_row], start=1):
        if cell.value:
            cell_value = str(cell.value).strip().lower()
            for target in target_columns:
                if target.lower().strip() == cell_value:
                    column_map[target] = col_idx
                    break

    return column_map


def find_item_row(ws, item_number: str, item_col_idx: int, start_row: int = 2) -> Optional[int]:
    """
    Find the row number for a given Item#.

    Args:
        ws: openpyxl worksheet
        item_number: The Item# to find
        item_col_idx: Column index containing Item#
        start_row: Row to start searching from (default: 2, after header)

    Returns:
        Row number (1-based) or None if not found
    """
    for row_idx in range(start_row, ws.max_row + 1):
        cell_value = ws.cell(row=row_idx, column=item_col_idx).value
        if cell_value is not None:
            if str(cell_value).strip() == str(item_number).strip():
                return row_idx
    return None


def _col_num_to_letter(col: int) -> str:
    """Convert column number (1-based) to Excel letter (A, B, ..., Z, AA, AB, ...)."""
    col_letter = ""
    temp_col = col
    while temp_col > 0:
        temp_col, remainder = divmod(temp_col - 1, 26)
        col_letter = chr(65 + remainder) + col_letter
    return col_letter


def _apply_colors_to_original(
    original_path: Path,
    output_path: Path,
    cell_styles: dict[tuple[int, int], str],
) -> None:
    """
    Apply cell background colors directly to original Excel, preserving everything else.

    This function preserves all original cell formatting by:
    1. Reading each cell's existing style index from the sheet XML
    2. Finding the complete xf element for that style in styles.xml
    3. Appending a copy with only the fillId changed
    4. Using a cache to avoid creating duplicate styles

    Uses careful string manipulation to preserve the original XML structure.

    Args:
        original_path: Path to the original Excel file
        output_path: Path for the output file
        cell_styles: Dict mapping (row, col) to color name ("green", "red", "yellow")
    """
    import re

    # Color hex values
    colors = {
        "green": "90EE90",
        "red": "FFB6B6",
        "yellow": "FFFACD",
    }

    # Copy original to output
    shutil.copy2(original_path, output_path)

    with tempfile.TemporaryDirectory() as tmpdir:
        tmpdir = Path(tmpdir)
        extract_dir = tmpdir / 'excel'

        # Extract the Excel file
        with zipfile.ZipFile(output_path, 'r') as z:
            z.extractall(extract_dir)

        # Read sheet XML to find each cell's current style
        sheet_path = extract_dir / 'xl' / 'worksheets' / 'sheet1.xml'
        with open(sheet_path, 'r', encoding='utf-8') as f:
            sheet_content = f.read()

        # Build a map of cell_ref -> current style index
        cell_current_styles: dict[str, int] = {}
        for (row, col), _ in cell_styles.items():
            cell_ref = f"{_col_num_to_letter(col)}{row}"
            style_match = re.search(rf'<c r="{cell_ref}"[^>]* s="(\d+)"', sheet_content)
            if style_match:
                cell_current_styles[cell_ref] = int(style_match.group(1))
            else:
                cell_current_styles[cell_ref] = 0

        # Read styles.xml
        styles_path = extract_dir / 'xl' / 'styles.xml'
        with open(styles_path, 'r', encoding='utf-8') as f:
            styles_content = f.read()

        # Step 1: Add our fill colors to the fills section
        # Find </fills> and insert before it
        fills_count_match = re.search(r'<fills count="(\d+)"', styles_content)
        if not fills_count_match:
            log_warning("Could not find fills count in styles.xml")
            return

        existing_fill_count = int(fills_count_match.group(1))
        fill_indices = {}

        new_fills = ""
        for i, (color_name, color_hex) in enumerate(colors.items()):
            fill_indices[color_name] = existing_fill_count + i
            new_fills += f'<fill><patternFill patternType="solid"><fgColor rgb="FF{color_hex}"/><bgColor indexed="64"/></patternFill></fill>'

        # Update fills count
        styles_content = styles_content.replace(
            f'<fills count="{existing_fill_count}"',
            f'<fills count="{existing_fill_count + 3}"'
        )
        # Insert new fills before </fills>
        styles_content = styles_content.replace('</fills>', f'{new_fills}</fills>')

        # Step 2: Extract all xf elements from cellXfs
        cellxfs_match = re.search(r'<cellXfs count="(\d+)">(.*?)</cellXfs>', styles_content, re.DOTALL)
        if not cellxfs_match:
            log_warning("Could not find cellXfs in styles.xml")
            return

        xf_count = int(cellxfs_match.group(1))
        cellxfs_content = cellxfs_match.group(2)

        # Extract individual xf elements (both self-closing and with children)
        # Pattern: self-closing <xf .../> OR with children <xf ...>...</xf>
        xf_elements = re.findall(r'<xf [^>]*/>|<xf [^>]*>.*?</xf>', cellxfs_content, re.DOTALL)

        # Cache: (original_style_idx, color_name) -> new_style_idx
        style_cache: dict[tuple[int, str], int] = {}
        new_xfs = ""

        for (row, col), color_name in cell_styles.items():
            cell_ref = f"{_col_num_to_letter(col)}{row}"
            orig_style_idx = cell_current_styles.get(cell_ref, 0)

            cache_key = (orig_style_idx, color_name)
            if cache_key not in style_cache:
                # Get the original xf element
                if orig_style_idx < len(xf_elements):
                    orig_xf = xf_elements[orig_style_idx]
                else:
                    orig_xf = '<xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/>'

                # Create new xf by modifying fillId
                new_fill_id = fill_indices[color_name]

                # Replace or add fillId attribute
                if 'fillId="' in orig_xf:
                    new_xf = re.sub(r'fillId="\d+"', f'fillId="{new_fill_id}"', orig_xf)
                else:
                    # Add fillId after xf tag opening
                    new_xf = orig_xf.replace('<xf ', f'<xf fillId="{new_fill_id}" ')

                # Ensure applyFill="1" is set
                if 'applyFill="' in new_xf:
                    new_xf = re.sub(r'applyFill="\d+"', 'applyFill="1"', new_xf)
                else:
                    new_xf = new_xf.replace('<xf ', '<xf applyFill="1" ')

                new_xfs += new_xf
                style_cache[cache_key] = xf_count
                xf_count += 1

        # Update cellXfs count and add new xf elements
        styles_content = styles_content.replace(
            f'<cellXfs count="{cellxfs_match.group(1)}"',
            f'<cellXfs count="{xf_count}"'
        )
        styles_content = styles_content.replace('</cellXfs>', f'{new_xfs}</cellXfs>')

        # Save modified styles.xml
        with open(styles_path, 'w', encoding='utf-8') as f:
            f.write(styles_content)

        # Step 3: Update sheet XML to apply new styles to cells
        for (row, col), color_name in cell_styles.items():
            cell_ref = f"{_col_num_to_letter(col)}{row}"
            orig_style_idx = cell_current_styles.get(cell_ref, 0)
            new_style_idx = style_cache[(orig_style_idx, color_name)]

            # Update or add the style reference
            cell_with_style = rf'(<c r="{cell_ref}"[^>]*) s="\d+"([^>]*>)'
            cell_without_style = rf'(<c r="{cell_ref}")([^>]*>)'

            if re.search(cell_with_style, sheet_content):
                sheet_content = re.sub(
                    cell_with_style,
                    rf'\1 s="{new_style_idx}"\2',
                    sheet_content
                )
            elif re.search(cell_without_style, sheet_content):
                sheet_content = re.sub(
                    cell_without_style,
                    rf'\1 s="{new_style_idx}"\2',
                    sheet_content
                )

        # Save modified sheet
        with open(sheet_path, 'w', encoding='utf-8') as f:
            f.write(sheet_content)

        # Repack the Excel file
        with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as z:
            for file_path in extract_dir.rglob('*'):
                if file_path.is_file():
                    arcname = file_path.relative_to(extract_dir).as_posix()
                    z.write(file_path, arcname)


def color_excel_cells(
    excel_path: str | Path,
    verification_results: dict[str, dict[str, bool]],
    output_path: Optional[str | Path] = None,
    columns_to_check: Optional[list[str]] = None,
) -> ColoringResult:
    """
    Color Excel cells based on verification results.

    Preserves images and other embedded content from the original file.

    Args:
        excel_path: Path to the Excel file
        verification_results: Dict mapping Item# -> {field_name: matched (bool)}
            Example: {"199034": {"EAN": True, "Name ENG": True, "Batch no:": False}}
        output_path: Where to save the colored Excel. If None, overwrites original.
        columns_to_check: List of column names to check. If None, uses all columns
                         from verification_results.

    Returns:
        ColoringResult with summary of changes made.
    """
    excel_path = Path(excel_path)

    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    # Determine save path
    save_path = Path(output_path) if output_path else excel_path

    # Load workbook just to read structure (we won't save via openpyxl)
    log_info(f"Loading Excel file: {excel_path.name}")
    wb = load_workbook(excel_path)
    ws = wb.active

    # Determine columns to process
    if columns_to_check is None:
        # Collect all unique field names from results
        all_fields = set()
        for fields in verification_results.values():
            all_fields.update(fields.keys())
        columns_to_check = ["Item#"] + list(all_fields)

    # Find column indices
    column_indices = find_column_indices(ws, columns_to_check)

    if "Item#" not in column_indices:
        raise ValueError("Could not find 'Item#' column in Excel file")

    item_col_idx = column_indices["Item#"]

    # Track results and cell styles to apply
    result = ColoringResult(
        excel_path=str(excel_path),
        products_found=0,
        products_not_found=0,
        cells_green=0,
        cells_red=0,
        cells_yellow=0,
        cell_details=[],
    )

    # Collect styles: {(row, col): color_name}
    cell_styles: dict[tuple[int, int], str] = {}

    # Process each item in verification results
    items_processed = set()
    for item_number, field_results in verification_results.items():
        row_idx = find_item_row(ws, item_number, item_col_idx)

        if row_idx is None:
            log_warning(f"Item# '{item_number}' not found in Excel")
            result.products_not_found += 1
            continue

        result.products_found += 1
        items_processed.add(item_number)

        # Color each field cell for this product
        for field_name, matched in field_results.items():
            if field_name not in column_indices:
                continue

            col_idx = column_indices[field_name]
            cell = ws.cell(row=row_idx, column=col_idx)
            cell_value = str(cell.value) if cell.value else ""

            if matched:
                cell_styles[(row_idx, col_idx)] = "green"
                result.cells_green += 1
                color = "green"
            else:
                cell_styles[(row_idx, col_idx)] = "red"
                result.cells_red += 1
                color = "red"

            result.cell_details.append(CellColorResult(
                row=row_idx,
                column=field_name,
                color=color,
                field_name=field_name,
                value=cell_value[:50] if cell_value else "",
            ))

        # Color unchecked columns (columns we have in Excel but weren't verified)
        for col_name, col_idx in column_indices.items():
            if col_name == "Item#":
                continue
            if col_name not in field_results:
                cell = ws.cell(row=row_idx, column=col_idx)
                # Only color if cell has a value
                if cell.value:
                    cell_styles[(row_idx, col_idx)] = "yellow"
                    result.cells_yellow += 1
                    result.cell_details.append(CellColorResult(
                        row=row_idx,
                        column=col_name,
                        color="yellow",
                        field_name=col_name,
                        value=str(cell.value)[:50] if cell.value else "",
                    ))

    # Close workbook without saving
    wb.close()

    # Apply colors directly to the original Excel file (preserves images)
    log_info(f"Applying colors to Excel (preserving images)...")
    _apply_colors_to_original(excel_path, save_path, cell_styles)

    log_info(f"Colored {result.cells_green} green, {result.cells_red} red, {result.cells_yellow} yellow cells")
    log_info(f"Saved to: {save_path.name}")

    return result
